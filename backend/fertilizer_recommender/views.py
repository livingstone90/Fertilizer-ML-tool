from django.shortcuts import render
import psycopg2
from django.contrib import messages
from openpyxl import load_workbook
from django_pandas.io import read_frame
from django.shortcuts import render, get_object_or_404
from rest_framework import generics
from django.http import JsonResponse
from django.views.generic import DetailView, View
from fertilizer_recommender.models import Data
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.views.decorators.csrf import csrf_exempt
import json

# Create your views here.


def get_villages(request):
    villages = Data.objects.values_list('village', flat=True).distinct()
    return JsonResponse(list(villages), safe=False)


def get_villages_data(request):
    villages = Data.objects.all().values('village', 'latitude', 'longitude')
    return JsonResponse(list(villages), safe=False)


# def village_predictions(request):
    # Query the database for all instances of Village
    #villages = Data.objects.all().distinct()

    # Create a list of dictionaries containing the data for each village
   # data = []
    # for village in villages:
    # data.append({
    # 'name': village.village,
    # 'latitude': float(village.latitude),
    # 'longitude': float(village.longitude),
    # 'altitude': float(village.altitude),
    # 'crop_yield_difference': village.yield_difference,
    # 'predicted_practice': village.predicted_practice
    # })

    # Return the data as JSON
    return JsonResponse({'data': data})


def village_predictions(request):
    # Query the database for unique instances of Village with non-null predicted practice
    villages = Data.objects.exclude(
        predicted_practice__isnull=True).values('village').distinct()

    # Create a list of dictionaries containing the data for each village
    data = []
    for village in villages:
        # Query the database for the data for this village
        village_data = Data.objects.filter(village=village['village']).exclude(
            predicted_practice__isnull=True)

        # Calculate the average values for each data field
        if village_data:
            latitude = sum([float(d.latitude)
                           for d in village_data]) / len(village_data)
            longitude = sum([float(d.longitude)
                            for d in village_data]) / len(village_data)
            altitude = sum([float(d.altitude)
                           for d in village_data]) / len(village_data)
            # Assumes all instances have the same predicted practice
            predicted_practice = village_data[0].predicted_practice

            # Add the data for this village to the list
            data.append({
                'name': village['village'],
                'latitude': latitude,
                'longitude': longitude,
                'altitude': altitude,
                'predicted_practice': predicted_practice
            })

    # Return the data as JSON
    return JsonResponse({'data': data})


class VillageDetailView(View):
    def get(self, request, village_name, *args, **kwargs):
        villages = Data.objects.filter(village=village_name)
        if not villages.exists():
            return JsonResponse({'error': 'Village not found'})
        village = villages.first()
        data = {
            'name': village.village,
            'longitude': village.longitude,
            'latitude': village.latitude,
            'altitude': village.altitude,
            'farmer_practice': village.farmer_practice,
            'fertilizer_practice': village.fertilizer_practice,
            'crop_yield_difference': village.yield_difference,
            'predicted_practice': village.predicted_practice,
        }
        return JsonResponse(data)


def upload_data(request):
    if request.method == 'POST':
        file = request.FILES['file']
        if not file.name.endswith('.xlsx'):
            messages.error(request, 'This is not an xlsx file')
        sheet_name = request.POST.get('sheet_name', None)
        wb = load_workbook(filename=file, read_only=True, data_only=True)
        ws = wb[sheet_name]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            data.append(Data(
                village=row[5],
                manure_applied=row[6],
                latitude=row[40],
                longitude=row[41],
                altitude=row[42],
                slope_dem=row[56],
                man_appllast_season=row[31],
                man_appl_2_season=row[32],
                man_appl_3_season=row[33],
                man_appl_4_season=row[34],
                man_appl_5_season=row[35],
                man_frequency=row[36],
                plant_date=row[49],
                land_form=row[55],
                slope=row[57],
                slope_position=row[58],
                wi=row[60],
                tpi=row[59],
                wealth_class=row[11],
                farmer_practice=row[66],
                fertilizer_practice=row[67],
                yield_difference=row[68]
                # add more columns as needed
            ))
        Data.objects.bulk_create(data)
        return render(request, 'success.html')
    return render(request, 'upload.html')
